"""
עמוד 1 — שוק הצמ"ה: נתחי שוק, מגמות ודגמים מובילים
"""

import io
import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from src.loader import load_annual, FILE1, FILE2

st.markdown("""
<style>
  .block-container { padding-top: 1.2rem; padding-bottom: 1rem; }
  div[data-testid="metric-container"] {
    background: white; border: 1px solid #DDE3EC;
    border-radius: 10px; padding: 12px 16px;
  }
  div[data-testid="metric-container"] label { font-size: 13px; }
  .upload-success { color: #27AE60; font-weight: 600; font-size: 13px; }
</style>
""", unsafe_allow_html=True)

ALL_YEARS = list(range(2018, 2027))

# ── One-time initialization of year checkbox state ────────────────────────────
# Must run before any sidebar block so segment/manufacturer changes never reset years.
for _y in ALL_YEARS:
    _yk = f"mkt_yr_{_y}"
    if _yk not in st.session_state:
        st.session_state[_yk] = (_y < 2026)   # default: 2018-2025 checked, 2026 off

# Tableau 10 — vibrant yet professional, distinguishable at a glance
COLORS = [
    "#4E79A7", "#F28E2B", "#E15759", "#76B7B2",
    "#59A14F", "#EDC948", "#B07AA1", "#FF9DA7",
    "#9C755F", "#BAB0AC",
]


# ── Data & file helpers ───────────────────────────────────────────────────────
@st.cache_data(show_spinner="טוען נתונים...")
def load_data() -> pd.DataFrame:
    return load_annual()


def detect_dest_file(data: bytes) -> tuple[Path, str]:
    raw = pd.read_excel(io.BytesIO(data), sheet_name=0, header=None, nrows=4)
    text = raw.to_string()
    if any(y in text for y in ["2026", "2025", "2024"]):
        return FILE1, "2022–2026"
    return FILE2, "2017–2022"


def save_file(data: bytes, dest: Path) -> bool:
    try:
        dest.parent.mkdir(parents=True, exist_ok=True)
        dest.write_bytes(data)
        return True
    except Exception:
        return False


# ── Sidebar — year checkboxes ─────────────────────────────────────────────────
with st.sidebar:
    st.markdown("#### 📅 בחירת שנים")
    _btn_col1, _btn_col2 = st.columns(2)
    if _btn_col1.button("בחר הכל", key="mkt_sel_all", use_container_width=True):
        for _y in ALL_YEARS:
            st.session_state[f"mkt_yr_{_y}"] = True
        st.rerun()
    if _btn_col2.button("נקה הכל", key="mkt_clr", use_container_width=True):
        for _y in ALL_YEARS:
            st.session_state[f"mkt_yr_{_y}"] = False
        st.rerun()

    selected_years = []
    cols3 = st.columns(3)
    for i, y in enumerate(ALL_YEARS):
        label = "2026◑" if y == 2026 else str(y)
        if cols3[i % 3].checkbox(label, key=f"mkt_yr_{y}"):
            selected_years.append(y)

    if not selected_years:
        st.warning("בחר לפחות שנה אחת")
        st.stop()

    st.markdown("---")
    st.markdown("#### 📂 עדכון נתונים")
    st.caption("התקופה תזוהה אוטומטית מתוך הקובץ")
    uploaded = st.file_uploader("קובץ Excel", type=["xlsx", "xls"],
                                label_visibility="collapsed")
    if uploaded:
        raw_bytes = bytes(uploaded.getbuffer())
        try:
            dest_file, period_label = detect_dest_file(raw_bytes)
            if save_file(raw_bytes, dest_file):
                load_data.clear()
                st.markdown(
                    f"<p class='upload-success'>✅ קובץ {period_label} נטען בהצלחה</p>",
                    unsafe_allow_html=True,
                )
                st.rerun()
            else:
                st.error("שגיאה בשמירת הקובץ")
        except Exception as e:
            st.error(f"שגיאה: {e}")

    st.markdown("---")
    st.caption("מקור: מרשם רכב ישראלי | כלים חדשים בלבד")


# ── Load & filter by years ────────────────────────────────────────────────────
df_all = load_data()
df = df_all[
    (df_all["condition"] == "new") &
    (df_all["year"].isin(selected_years)) &
    (~df_all["is_aggregate"])
].copy()

# ── Segment list ──────────────────────────────────────────────────────────────
seg_df_raw = (
    df.groupby(["category_en", "category_he"])["units"]
    .sum().reset_index()
    .sort_values("units", ascending=False)
    .query("units > 0")
)
seg_labels  = [r.category_he or r.category_en for _, r in seg_df_raw.iterrows()]
seg_en_list = seg_df_raw["category_en"].tolist()

# ── Page header ───────────────────────────────────────────────────────────────
yr_str = f"{min(selected_years)}–{max(selected_years)}"
st.title("🏗️  שוק ציוד מכני הנדסי — ישראל")
st.caption(f"כלים חדשים בלבד | {yr_str}")

# ── Segment selector (persists across sidebar changes) ────────────────────────
col_seg, _ = st.columns([2, 3])
with col_seg:
    prev = st.session_state.get("_mkt_cat_en", "")
    dflt = seg_en_list.index(prev) if prev in seg_en_list else 0
    sel_idx = st.selectbox(
        "בחר סגמנט", range(len(seg_labels)),
        format_func=lambda i: seg_labels[i], index=dflt,
    )
    st.session_state["_mkt_cat_en"] = seg_en_list[sel_idx]

sel_cat_en = seg_en_list[sel_idx]
sel_cat_he = seg_df_raw.iloc[sel_idx]["category_he"] or sel_cat_en
st.markdown("---")

# ── Segment data ──────────────────────────────────────────────────────────────
seg = df[df["category_en"] == sel_cat_en].copy()
seg_total = int(seg["units"].sum())

mfr_tbl = (
    seg.groupby("manufacturer_en")["units"]
    .sum().reset_index()
    .sort_values("units", ascending=False)
)
mfr_tbl["share"] = (mfr_tbl["units"] / seg_total * 100).round(1)
all_mfrs = mfr_tbl["manufacturer_en"].tolist()

# ── Sidebar — manufacturer multiselect (segment-aware) ───────────────────────
with st.sidebar:
    st.markdown("---")
    st.markdown("#### 🏭 יצרנים לתצוגה")
    # Reset selection when segment changes
    if st.session_state.get("_mkt_seg") != sel_cat_en:
        st.session_state["_mkt_seg"] = sel_cat_en
        st.session_state["_mkt_mfrs"] = all_mfrs[:min(5, len(all_mfrs))]

    # Filter any previously stored selection to only manufacturers in this segment
    # (avoids crash when stored names like 'DOOSAN' don't exist in new segment)
    stored = st.session_state.get("_mkt_mfrs", [])
    valid_default = [m for m in stored if m in all_mfrs] or all_mfrs[:min(5, len(all_mfrs))]

    selected_mfrs = st.multiselect(
        "יצרנים",
        all_mfrs,
        default=valid_default,
        label_visibility="collapsed",
    )
    st.session_state["_mkt_mfrs"] = selected_mfrs   # persist without key binding

    if not selected_mfrs:
        st.warning("בחר לפחות יצרן אחד")
        st.stop()

# ── Color map (consistent across all charts) ──────────────────────────────────
color_map = {m: COLORS[i % len(COLORS)] for i, m in enumerate(all_mfrs)}
color_map["אחרים"] = "#C8C8C8"


# ── Excel report generator ────────────────────────────────────────────────────
@st.cache_data(show_spinner=False)
def build_excel(seg_df: pd.DataFrame, mfr_df: pd.DataFrame,
                cat_he: str, yr_str: str) -> bytes:
    BLUE, LIGHT = "1B4F91", "EBF3FB"

    def hdr(ws, row, labels):
        for col, lbl in enumerate(labels, 1):
            c = ws.cell(row=row, column=col, value=lbl)
            c.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
            c.fill = PatternFill("solid", fgColor=BLUE)
            c.alignment = Alignment(horizontal="center", vertical="center")

    def stripe(ws, row, values):
        for col, val in enumerate(values, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = Font(name="Arial", size=10)
            if row % 2 == 0:
                c.fill = PatternFill("solid", fgColor=LIGHT)

    def widths(ws, col_widths):
        for col, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w

    wb = Workbook()

    # ── Sheet 1: יצרנים ──────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "יצרנים"
    ws1.merge_cells("A1:G1")
    t = ws1["A1"]
    t.value = f"{cat_he} | {yr_str}"
    t.font = Font(bold=True, name="Arial", size=14, color=BLUE)
    t.alignment = Alignment(horizontal="center")
    ws1.row_dimensions[1].height = 26

    hdr(ws1, 3, ["יצרן", "יחידות", "נתח שוק", "דגם #1", "יח' #1", "דגם #2", "יח' #2"])
    ws1.row_dimensions[3].height = 20

    for r, (_, row) in enumerate(mfr_df.iterrows(), 4):
        m = row["manufacturer_en"]
        top = (seg_df[seg_df["manufacturer_en"] == m]
               .groupby("model_name")["units"].sum()
               .sort_values(ascending=False))
        m1n = top.index[0] if len(top) >= 1 else "—"
        m1v = int(top.iloc[0]) if len(top) >= 1 else 0
        m2n = top.index[1] if len(top) >= 2 else "—"
        m2v = int(top.iloc[1]) if len(top) >= 2 else 0
        stripe(ws1, r, [m, int(row["units"]), f"{row['share']}%", m1n, m1v, m2n, m2v])

    widths(ws1, [22, 10, 10, 20, 10, 20, 10])

    # ── Sheet 2: מגמות שנתיות ────────────────────────────────────────────────
    ws2 = wb.create_sheet("מגמות שנתיות")
    pivot = (
        seg_df.groupby(["manufacturer_en", "year"])["units"].sum()
        .unstack(fill_value=0)
        .sort_values(by=sorted(seg_df["year"].unique())[-1], ascending=False)
    )
    hdr(ws2, 1, ["יצרן"] + [str(y) for y in pivot.columns])
    for r, (mfr_name, row) in enumerate(pivot.iterrows(), 2):
        stripe(ws2, r, [mfr_name] + [int(v) for v in row.values])
    widths(ws2, [22] + [11] * len(pivot.columns))

    # ── Sheet 3: דגמים מובילים ───────────────────────────────────────────────
    ws3 = wb.create_sheet("דגמים מובילים")
    top_mdl = (
        seg_df.groupby(["manufacturer_en", "model_name"])["units"]
        .sum().reset_index().sort_values("units", ascending=False).head(25)
    )
    hdr(ws3, 1, ["יצרן", "דגם", "יחידות", "% מסגמנט"])
    seg_tot = int(seg_df["units"].sum())
    for r, (_, row) in enumerate(top_mdl.iterrows(), 2):
        pct = round(row["units"] / seg_tot * 100, 1) if seg_tot else 0
        stripe(ws3, r, [row["manufacturer_en"], row["model_name"],
                        int(row["units"]), f"{pct}%"])
    widths(ws3, [22, 20, 12, 12])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── KPIs ──────────────────────────────────────────────────────────────────────
rank1 = mfr_tbl.iloc[0] if len(mfr_tbl) >= 1 else None
rank2 = mfr_tbl.iloc[1] if len(mfr_tbl) >= 2 else None

annual      = seg.groupby("year")["units"].sum()
annual_full = annual[annual.index <= 2025]
yoy_label, yoy_str = "שינוי שנתי", "—"
if len(annual_full) >= 2:
    yr_last  = sorted(annual_full.index)[-1]
    yr_prev  = sorted(annual_full.index)[-2]
    delta    = int(annual_full[yr_last]) - int(annual_full[yr_prev])
    pct      = delta / annual_full[yr_prev] * 100 if annual_full[yr_prev] else 0
    yoy_label = f"שינוי {yr_prev}→{yr_last}"
    yoy_str   = f"{delta:+,} ({pct:+.1f}%)"

# HHI — market concentration
hhi_val = int((mfr_tbl["share"] ** 2).sum())
if hhi_val < 1_500:
    hhi_lbl, hhi_clr = "שוק תחרותי 🟢", "#27AE60"
elif hhi_val < 2_500:
    hhi_lbl, hhi_clr = "ריכוז מתון 🟡",  "#E67E22"
else:
    hhi_lbl, hhi_clr = "שוק מרוכז 🔴",  "#E74C3C"

k1, k2, k3, k4, k5 = st.columns(5)
k1.metric("סה\"כ יחידות", f"{seg_total:,}")
k2.metric("מוביל שוק 🥇",
          rank1.manufacturer_en if rank1 is not None else "—",
          f"{rank1.share}%" if rank1 is not None else "")
k3.metric("שני בשוק 🥈",
          rank2.manufacturer_en if rank2 is not None else "—",
          f"{rank2.share}%" if rank2 is not None else "")
k4.metric(yoy_label, yoy_str)

with k5:
    st.markdown(
        f"""<div style="background:white;border:1px solid #DDE3EC;border-radius:10px;
            padding:12px 16px;height:100%;">
          <div style="font-size:13px;color:#555;margin-bottom:4px;">ריכוזיות שוק (HHI)</div>
          <div style="font-size:26px;font-weight:700;color:#1A1A2E;line-height:1.1;">{hhi_val:,}</div>
          <div style="font-size:13px;font-weight:600;color:{hhi_clr};margin-top:4px;">{hhi_lbl}</div>
        </div>""",
        unsafe_allow_html=True,
    )

# ── Excel download (right-aligned, after KPIs) ───────────────────────────────
st.markdown("<br>", unsafe_allow_html=True)
_, dl_col = st.columns([4, 1])
with dl_col:
    excel_bytes = build_excel(seg, mfr_tbl, sel_cat_he, yr_str)
    st.download_button(
        "📥 הורד דוח Excel",
        data=excel_bytes,
        file_name=f"market_{sel_cat_en}_{yr_str}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

# ── Top 2 models per selected manufacturer ────────────────────────────────────
st.markdown("<br>", unsafe_allow_html=True)
st.markdown("#### 🏆 יצרנים מובילים — דגמי דגל")

rows = []
for _, r in mfr_tbl[mfr_tbl["manufacturer_en"].isin(selected_mfrs)].iterrows():
    m = r["manufacturer_en"]
    top = (
        seg[seg["manufacturer_en"] == m]
        .groupby("model_name")["units"].sum()
        .sort_values(ascending=False)
    )
    m1 = f"{top.index[0]} ({int(top.iloc[0]):,})" if len(top) >= 1 else "—"
    m2 = f"{top.index[1]} ({int(top.iloc[1]):,})" if len(top) >= 2 else "—"
    rows.append({"יצרן": m, "נתח שוק": f"{r['share']}%", "דגם #1": m1, "דגם #2": m2})

st.dataframe(pd.DataFrame(rows), hide_index=True, use_container_width=True)

st.markdown("---")

# ── Row 1: Donut + Grouped trend bar ─────────────────────────────────────────
col_pie, col_bar = st.columns([1, 1.6])

with col_pie:
    st.markdown(f"#### נתחי שוק — {sel_cat_he}")
    top_pie = mfr_tbl[mfr_tbl["manufacturer_en"].isin(selected_mfrs)].copy()
    others_u = mfr_tbl[~mfr_tbl["manufacturer_en"].isin(selected_mfrs)]["units"].sum()
    if others_u > 0:
        top_pie = pd.concat([top_pie, pd.DataFrame([{
            "manufacturer_en": "אחרים", "units": others_u,
            "share": round(others_u / seg_total * 100, 1),
        }])], ignore_index=True)

    fig_pie = px.pie(
        top_pie, names="manufacturer_en", values="units", hole=0.42,
        color="manufacturer_en",
        color_discrete_map=color_map,
    )
    fig_pie.update_traces(
        textinfo="percent+label", textfont_size=12,
        pull=[0.04] + [0] * (len(top_pie) - 1),
    )
    fig_pie.update_layout(
        height=370, showlegend=False,
        margin=dict(t=10, b=10, l=10, r=10),
        paper_bgcolor="rgba(0,0,0,0)",
    )
    st.plotly_chart(fig_pie, use_container_width=True, config={"displayModeBar": False})

with col_bar:
    st.markdown("#### מגמה שנתית לפי יצרן")
    trend = (
        seg[seg["manufacturer_en"].isin(selected_mfrs)]
        .groupby(["year", "manufacturer_en"])["units"].sum().reset_index()
    )
    full_idx = pd.MultiIndex.from_product(
        [selected_years, selected_mfrs], names=["year", "manufacturer_en"]
    )
    trend = (
        trend.set_index(["year", "manufacturer_en"])
        .reindex(full_idx, fill_value=0)
        .reset_index()
    )

    fig_trend = go.Figure()
    for m in selected_mfrs:
        d = trend[trend["manufacturer_en"] == m].sort_values("year")
        fig_trend.add_trace(go.Bar(
            name=m, x=d["year"], y=d["units"],
            marker_color=color_map.get(m, "#888"),
            text=d["units"].where(d["units"] > 0, ""),
            textposition="outside", textfont_size=9,
        ))
    fig_trend.update_layout(
        barmode="group", height=370,
        xaxis=dict(dtick=1, tickformat="d", title=""),
        yaxis=dict(title="יחידות"),
        legend=dict(title="יצרן", orientation="v", x=1.01),
        plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
        margin=dict(t=10, b=30, l=0, r=10),
    )
    fig_trend.update_xaxes(showgrid=False)
    fig_trend.update_yaxes(showgrid=True, gridcolor="#EEE")
    st.plotly_chart(fig_trend, use_container_width=True, config={"displayModeBar": False})

# ── Market share % evolution (stacked 100% bar) ───────────────────────────────
st.markdown("#### 📈 התפתחות נתחי שוק לאורך השנים")

# Compute shares from raw units (not from summed pre-rounded floats) to avoid
# floating-point artefacts like 36.300000000000004.
share_units = seg.groupby(["year", "manufacturer_en"])["units"].sum().reset_index()
total_by_yr  = seg.groupby("year")["units"].sum()

sel_share = share_units[share_units["manufacturer_en"].isin(selected_mfrs)].copy()
sel_share["pct"] = sel_share.apply(
    lambda r: round(r["units"] / total_by_yr[r["year"]] * 100, 1)
    if total_by_yr.get(r["year"], 0) > 0 else 0.0, axis=1
)

oth_units = (
    share_units[~share_units["manufacturer_en"].isin(selected_mfrs)]
    .groupby("year")["units"].sum().reset_index()
)
oth_units["manufacturer_en"] = "אחרים"
oth_units["pct"] = oth_units.apply(
    lambda r: round(r["units"] / total_by_yr[r["year"]] * 100, 1)
    if total_by_yr.get(r["year"], 0) > 0 else 0.0, axis=1
)

share_plot = pd.concat([sel_share, oth_units], ignore_index=True)

fig_evo = go.Figure()
has_others = oth_units["pct"].sum() > 0
plot_order = selected_mfrs + (["אחרים"] if has_others else [])
for m in plot_order:
    d = share_plot[share_plot["manufacturer_en"] == m].sort_values("year")
    # Only show label when the slice is wide enough to be readable (≥ 6 %)
    labels = d["pct"].apply(lambda x: f"{x:.1f}%" if x >= 6 else "")
    fig_evo.add_trace(go.Bar(
        name=m, x=d["year"], y=d["pct"],
        marker_color=color_map.get(m, "#888"),
        text=labels,
        textposition="inside",
        insidetextanchor="middle",
        textfont=dict(size=12, color="white"),
        constraintext="none",
    ))
fig_evo.update_layout(
    barmode="stack", height=400,
    xaxis=dict(dtick=1, tickformat="d", title=""),
    yaxis=dict(title="נתח שוק (%)", range=[0, 100]),
    legend=dict(title="יצרן", orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
    plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
    margin=dict(t=40, b=20, l=0, r=10),
)
fig_evo.update_xaxes(showgrid=False)
fig_evo.update_yaxes(showgrid=True, gridcolor="#EEE")
st.plotly_chart(fig_evo, use_container_width=True, config={"displayModeBar": False})

# ── Top 15 models horizontal bar ──────────────────────────────────────────────
st.markdown("#### הדגמים המובילים")
models = (
    seg.groupby(["manufacturer_en", "model_name"])["units"]
    .sum().reset_index()
    .sort_values("units", ascending=False)
    .head(15).copy()
)
models["תווית"] = models["manufacturer_en"] + "  –  " + models["model_name"].fillna("?")

fig_mdl = px.bar(
    models[::-1], x="units", y="תווית", orientation="h",
    text="units", color="units",
    color_continuous_scale=["#AED6F1", "#1B4F91"],
    labels={"units": "יחידות", "תווית": ""},
)
fig_mdl.update_traces(textposition="outside", textfont_size=11)
fig_mdl.update_layout(
    height=max(280, len(models) * 30),
    showlegend=False, coloraxis_showscale=False,
    plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
    margin=dict(t=10, b=10, l=10, r=60),
    xaxis=dict(showgrid=True, gridcolor="#EEE"),
    yaxis=dict(showgrid=False),
)
st.plotly_chart(fig_mdl, use_container_width=True, config={"displayModeBar": False})
